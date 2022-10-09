from osgeo import gdal

def calculate():
    dataset = gdal.Open("C:\czg/602\project/Uav path planning/1.code\pycharm code\photo_to_gps1.0\dsm.tif")  # 打开tif
    geo_information = dataset.GetGeoTransform()
    print(geo_information)
    col = dataset.RasterXSize  # 行数
    row = dataset.RasterYSize  # 列数
    # band = dataset.RasterCount  # 波段

    top_left_corner = [0, 0]  # 左上角
    bottom_left_corner = [col, 0]  # 左下角
    top_right_corner = [0, row]  # 右上角
    bottom_right_corner = [col, row]  # 右下角

    # 左上角经纬度
    top_left_corner_lon = geo_information[0] + top_left_corner[0] * geo_information[1] + top_left_corner[1] * \
                          geo_information[2]
    top_left_corner_lat = geo_information[3] + top_left_corner[0] * geo_information[4] + top_left_corner[1] * \
                          geo_information[5]

    # 左下角经纬度
    bottom_left_corner_lon = geo_information[0] + bottom_left_corner[0] * geo_information[1] + bottom_left_corner[1] * \
                             geo_information[2]
    bottom_left_corner_lat = geo_information[3] + bottom_left_corner[0] * geo_information[4] + bottom_left_corner[1] * \
                             geo_information[5]

    # 右上角经纬度
    top_right_corner_lon = geo_information[0] + top_right_corner[0] * geo_information[1] + top_right_corner[1] * \
                           geo_information[2]
    top_right_corner_lat = geo_information[3] + top_right_corner[0] * geo_information[4] + top_right_corner[1] * \
                           geo_information[5]

    # 右下角经纬度
    bottom_right_corner_lon = geo_information[0] + bottom_right_corner[0] * geo_information[1] + bottom_right_corner[
        1] * geo_information[2]
    bottom_right_corner_lat = geo_information[3] + bottom_right_corner[0] * geo_information[4] + bottom_right_corner[
        1] * geo_information[5]
    return top_left_corner_lon, top_left_corner_lat, bottom_left_corner_lon, bottom_left_corner_lat, top_right_corner_lon, top_right_corner_lat, bottom_right_corner_lon, bottom_right_corner_lat


from osgeo import gdal
from pylab import *  # 支持中文
mpl.rcParams['font.sans-serif'] = ['SimHei']
from openpyxl import Workbook
work = Workbook()

def out(data, name):
    ws = work.active
    ws['A1'] = '经度'
    ws['B1'] = '纬度'
    ws['C1'] = '高程'
    ws['D1'] = '所在栅格行'
    ws['E1'] = '所在栅格列'
    for i in range(len(data)):
        rows = []
        row_length = len(data[i])
        if row_length != 0:
            for j in range(row_length):
                rows.append(data[i][j])
                ws.append(rows[j])
        # print(rows)
    work.save(name)


def get_tiff_gps():
    dataset = gdal.Open("C:\czg/602\project/Uav path planning/1.code\pycharm code\photo_to_gps1.0\dsm.tif")

    # 获取行数列数和地理信息
    # geo_information(0):左上像素左上角的x坐标。
    # geo_information(1):w - e像素分辨率 / 像素宽度。
    # geo_information(2):行旋转（通常为零）。
    # geo_information(3):左上像素左上角的y坐标。
    # geo_information(4):列旋转（通常为零）。
    # geo_information(5):n - s像素分辨率 / 像素高度（北半球上图像为负值）
    geo_information = dataset.GetGeoTransform()
    col = dataset.RasterXSize  # 438
    row = dataset.RasterYSize  # 671
    band = dataset.RasterCount
    dem = dataset.GetRasterBand(1).ReadAsArray()
    # 获取行列数，对应其经纬度,j对于x坐标
    cols = []
    for y in range(row):  # 行
        rows = []
        for x in range(col):  # 列
            # 有效高程
            if dem[y][x] > 0:
                # 输出经纬度
                lon = geo_information[0] + x * geo_information[1] + y * geo_information[2]
                lat = geo_information[3] + x * geo_information[4] + y * geo_information[5]
                child = [lon, lat, dem[y][x], y, x]
                rows.append(child)
        cols.append(rows)
    out(cols, 'C:\czg/602\project/Uav path planning/1.code\pycharm code\photo_to_gps1.0/1.xlsx')
    print('表已经生成')

print(calculate())
# get_tiff_gps()